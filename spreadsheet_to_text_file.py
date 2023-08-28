#!/usr/bin/env python3
# spreadsheetToTextFile.py — An exercise in manipulating Excel files.
# For more information, see README.md

import logging
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)
# logging.disable(logging.CRITICAL)

file_name = f'{input("Please enter file name here: ")}'
plus_extension = f"{file_name}.xlsx"


def open_workbook(file):
    """Open user-designated Excel file in current directory and return its sheet contents."""
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook.active
    return sheet


def record_contents(sheet):
    """Record contents of rows per column of sheet into a list of lists and return list."""
    return [
        [
            sheet[f"{get_column_letter(column + 1)}{row}"].value
            for row in range(1, sheet.max_row + 1)
            if sheet[f"{get_column_letter(column + 1)}{row}"].value is not None
        ]
        for column in range(sheet.max_column)
    ]


def write_contents(contents):
    """Write contents of each line of secondary list to file associated with primary list value."""
    for index_1, _ in enumerate(contents):
        with open(
            f"{file_name}_column{str(index_1 + 1).zfill(3)}.txt", "w", encoding="utf-8"
        ) as txt:
            for index_2, _ in enumerate(contents[index_1]):
                txt.write(str(contents[index_1][index_2]))


def main():
    """Runs prior functions in sequence."""
    user_sheet = open_workbook(plus_extension)
    file_contents = record_contents(user_sheet)
    write_contents(file_contents)


if __name__ == "__main__":
    main()
