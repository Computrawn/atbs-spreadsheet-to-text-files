#! python3
# spreadsheetToTextFile.py — An exercise in manipulating Excel files.
# For more information, see project_details.txt

import logging
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)
# logging.disable(logging.CRITICAL)


def open_workbook():
    """Open user-designated Excel file in current directory and return its sheet contents."""
    file_name = f'{input("Please enter file name here: ")}.xlsx'
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet = workbook.active
    return sheet


def record_contents(sheet):
    """Record contents of rows per column of sheet into a list of lists and return list."""
    file_contents = []

    for column in range(sheet.max_column):
        sheet_contents = []
        column_letter = get_column_letter(column + 1)

        for row in range(1, sheet.max_row + 1):
            cell_value = sheet[f"{column_letter}{row}"].value

            if cell_value is not None:
                sheet_contents.append(cell_value)

        file_contents.append(sheet_contents)
    return file_contents


def write_contents(contents):
    """Write contents of each line of secondary list to file associated with primary list value."""
    for index_1, _ in enumerate(contents):
        with open(
            f"Text_file_{str(index_1 + 1).zfill(3)}.txt", "w", encoding="utf-8"
        ) as txt:
            for index_2, _ in enumerate(contents[index_1]):
                txt.write(contents[index_1][index_2])


def main_func():
    """Runs prior functions in sequence."""
    user_sheet = open_workbook()
    file_contents = record_contents(user_sheet)
    write_contents(file_contents)


main_func()
